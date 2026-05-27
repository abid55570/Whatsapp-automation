"""GSTR-3B summary xlsx — owner uses this to fill the form on gst.gov.in.

GSTR-3B is the monthly summary return. We pre-fill sections:
    3.1  Outward + RCM supplies (taxable, IGST, CGST, SGST, cess)
    3.2  Inter-state to unregistered (B2CL split by state of supply)
    4    Eligible ITC (from purchase invoices)
    5    Exempt, nil-rated, non-GST inward supplies (manual; we leave blank)
    6.1  Payment of tax (computed from 3.1 + 4)
"""
from __future__ import annotations

import io
from collections import defaultdict
from datetime import date
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.models.enums import InvoiceStatus, InvoiceType, PurchaseInvoiceStatus

if TYPE_CHECKING:
    from app.models import Business, Invoice, PurchaseInvoice


_TITLE_FONT = Font(bold=True, size=14, color="0F172A")
_SECTION_FONT = Font(bold=True, size=11, color="FFFFFF")
_SECTION_FILL = PatternFill("solid", fgColor="16A34A")
_LABEL_FONT = Font(bold=True)
_RUPEE_FORMAT = '#,##0.00;[Red]-#,##0.00'


def _p(paise: int) -> float:
    return round((paise or 0) / 100, 2)


def _section_header(ws, row: int, label: str, span: int = 6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = _SECTION_FONT
    cell.fill = _SECTION_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22


def build_gstr3b_summary(
    business: "Business",
    invoices: list["Invoice"],
    purchase_invoices: list["PurchaseInvoice"],
    period_month: date,
) -> bytes:
    """Build the GSTR-3B summary xlsx."""
    active_inv = [i for i in invoices if i.status != InvoiceStatus.CANCELLED]
    active_purchases = [
        p for p in purchase_invoices if p.status != PurchaseInvoiceStatus.CANCELLED
    ]

    # ---- Aggregate 3.1: outward supplies ----
    # (a) Outward taxable (other than zero-rated, nil-rated, exempted)
    taxable_outward = sum(i.taxable_paise for i in active_inv if i.invoice_type != InvoiceType.EXPORT)
    igst_outward = sum(i.igst_paise for i in active_inv)
    cgst_outward = sum(i.cgst_paise for i in active_inv)
    sgst_outward = sum(i.sgst_paise for i in active_inv)
    cess_outward = sum(i.cess_paise for i in active_inv)

    # ---- 3.2: inter-state supplies to unregistered (B2CL) ----
    b2cl_by_pos: dict[str, dict] = defaultdict(lambda: {"txval": 0, "iamt": 0})
    for inv in active_inv:
        if inv.invoice_type == InvoiceType.B2C_LARGE:
            pos = inv.cx_state_code or "00"
            b2cl_by_pos[pos]["txval"] += inv.taxable_paise
            b2cl_by_pos[pos]["iamt"] += inv.igst_paise

    # ---- 4: ITC available from purchases ----
    eligible = [p for p in active_purchases if p.is_itc_eligible]
    capital = [p for p in eligible if p.is_capital_goods]
    inputs = [p for p in eligible if not p.is_capital_goods]

    def _sum(rows, attr):
        return sum(getattr(r, attr) for r in rows)

    itc_input_igst = _sum(inputs, "igst_paise")
    itc_input_cgst = _sum(inputs, "cgst_paise")
    itc_input_sgst = _sum(inputs, "sgst_paise")
    itc_input_cess = _sum(inputs, "cess_paise")
    itc_cap_igst = _sum(capital, "igst_paise")
    itc_cap_cgst = _sum(capital, "cgst_paise")
    itc_cap_sgst = _sum(capital, "sgst_paise")
    itc_cap_cess = _sum(capital, "cess_paise")

    # ---- 6.1: net tax payable ----
    net_igst = max(0, igst_outward - itc_input_igst - itc_cap_igst)
    net_cgst = max(0, cgst_outward - itc_input_cgst - itc_cap_cgst)
    net_sgst = max(0, sgst_outward - itc_input_sgst - itc_cap_sgst)
    net_cess = max(0, cess_outward - itc_input_cess - itc_cap_cess)

    # ============================================================
    # Build workbook
    # ============================================================
    wb = Workbook()
    ws = wb.active
    ws.title = "GSTR-3B Summary"

    # Title block
    ws.cell(row=1, column=1, value=f"GSTR-3B Summary — {business.name}").font = _TITLE_FONT
    ws.cell(row=2, column=1, value=f"GSTIN: {business.gstin or 'Not registered'}")
    ws.cell(
        row=3, column=1,
        value=f"Period: {period_month.strftime('%B %Y')}",
    )
    ws.cell(row=4, column=1, value=f"Invoices: {len(active_inv)} · Purchase bills: {len(active_purchases)}")

    r = 6
    _section_header(ws, r, "3.1 Details of Outward Supplies and Inward Supplies liable to Reverse Charge")
    r += 1
    headers = ["Nature", "Taxable Value (₹)", "IGST (₹)", "CGST (₹)", "SGST/UTGST (₹)", "Cess (₹)"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = _LABEL_FONT
        cell.alignment = Alignment(horizontal="center")
    r += 1
    rows_31 = [
        ("(a) Outward taxable supplies (other than zero-rated, nil-rated, exempted)",
         taxable_outward, igst_outward, cgst_outward, sgst_outward, cess_outward),
        ("(b) Outward taxable supplies (zero-rated)", 0, 0, 0, 0, 0),
        ("(c) Other outward supplies (nil-rated, exempted)", 0, 0, 0, 0, 0),
        ("(d) Inward supplies (liable to reverse charge)", 0, 0, 0, 0, 0),
        ("(e) Non-GST outward supplies", 0, 0, 0, 0, 0),
    ]
    for label, txval, igst, cgst, sgst, cess in rows_31:
        ws.cell(row=r, column=1, value=label)
        for c, paise in enumerate([txval, igst, cgst, sgst, cess], start=2):
            cell = ws.cell(row=r, column=c, value=_p(paise))
            cell.number_format = _RUPEE_FORMAT
            cell.alignment = Alignment(horizontal="right")
        r += 1

    r += 2
    _section_header(ws, r, "3.2 Inter-state supplies to unregistered persons / composition dealers")
    r += 1
    ws.cell(row=r, column=1, value="Place of Supply (State)").font = _LABEL_FONT
    ws.cell(row=r, column=2, value="Taxable Value (₹)").font = _LABEL_FONT
    ws.cell(row=r, column=3, value="IGST (₹)").font = _LABEL_FONT
    r += 1
    if not b2cl_by_pos:
        ws.cell(row=r, column=1, value="(no inter-state B2C-large supplies)").font = Font(italic=True)
        r += 1
    else:
        for pos, b in sorted(b2cl_by_pos.items()):
            ws.cell(row=r, column=1, value=f"State code {pos}")
            ws.cell(row=r, column=2, value=_p(b["txval"])).number_format = _RUPEE_FORMAT
            ws.cell(row=r, column=3, value=_p(b["iamt"])).number_format = _RUPEE_FORMAT
            r += 1

    r += 2
    _section_header(ws, r, "4. Eligible ITC")
    r += 1
    itc_headers = ["Particulars", "IGST (₹)", "CGST (₹)", "SGST (₹)", "Cess (₹)"]
    for c, h in enumerate(itc_headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = _LABEL_FONT
    r += 1
    rows_4 = [
        ("(A) ITC Available", itc_input_igst + itc_cap_igst, itc_input_cgst + itc_cap_cgst,
         itc_input_sgst + itc_cap_sgst, itc_input_cess + itc_cap_cess),
        ("    (1) Import of goods", 0, 0, 0, 0),
        ("    (2) Import of services", 0, 0, 0, 0),
        ("    (3) Inward supplies liable to reverse charge", 0, 0, 0, 0),
        ("    (4) Inward supplies from ISD", 0, 0, 0, 0),
        ("    (5) All other ITC (Inputs)", itc_input_igst, itc_input_cgst, itc_input_sgst, itc_input_cess),
        ("    (5a) Capital goods", itc_cap_igst, itc_cap_cgst, itc_cap_sgst, itc_cap_cess),
    ]
    for label, igst, cgst, sgst, cess in rows_4:
        ws.cell(row=r, column=1, value=label)
        for c, paise in enumerate([igst, cgst, sgst, cess], start=2):
            cell = ws.cell(row=r, column=c, value=_p(paise))
            cell.number_format = _RUPEE_FORMAT
            cell.alignment = Alignment(horizontal="right")
        r += 1

    r += 2
    _section_header(ws, r, "6.1 Payment of Tax")
    r += 1
    pay_headers = ["Particulars", "IGST (₹)", "CGST (₹)", "SGST (₹)", "Cess (₹)"]
    for c, h in enumerate(pay_headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = _LABEL_FONT
    r += 1
    rows_61 = [
        ("Tax payable (from 3.1)", igst_outward, cgst_outward, sgst_outward, cess_outward),
        ("ITC available (from 4A)", itc_input_igst + itc_cap_igst, itc_input_cgst + itc_cap_cgst,
         itc_input_sgst + itc_cap_sgst, itc_input_cess + itc_cap_cess),
        ("Tax to be paid in cash", net_igst, net_cgst, net_sgst, net_cess),
    ]
    for label, igst, cgst, sgst, cess in rows_61:
        ws.cell(row=r, column=1, value=label).font = (
            Font(bold=True) if "paid in cash" in label else Font()
        )
        for c, paise in enumerate([igst, cgst, sgst, cess], start=2):
            cell = ws.cell(row=r, column=c, value=_p(paise))
            cell.number_format = _RUPEE_FORMAT
            cell.font = Font(bold=True) if "paid in cash" in label else Font()
            cell.alignment = Alignment(horizontal="right")
        r += 1

    r += 2
    ws.cell(
        row=r, column=1,
        value="⚠️ Verify all figures with your CA before filing GSTR-3B on gst.gov.in. This summary is auto-computed from invoices issued via WhatsApp Auto.",
    ).font = Font(italic=True, color="DC2626")

    # Column widths
    for col, width in enumerate([54, 18, 16, 16, 18, 14], start=1):
        ws.column_dimensions[chr(64 + col)].width = width

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
