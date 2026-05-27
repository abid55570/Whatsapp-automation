"""ITR-4 (Sugam) Profit & Loss xlsx — annual filing for presumptive small biz.

Section 44AD (goods) and 44ADA (services) of the IT Act:
  - Eligible if FY turnover < ₹2 crore (goods) or ₹50 lakh (services).
  - Presumptive profit = 8% of cash turnover + 6% of digital turnover.
  - Optionally declare higher actual profit.

Output sheets:
  1. Overview          — FY label, turnover, presumed profit, tax liability hint
  2. Quarterly         — Q1–Q4 turnover & GST collected
  3. Income            — sales by month
  4. Expenses          — purchase totals by month + category
  5. Filing Worksheet  — pre-filled lines mapped to ITR-4 schedule names
"""
from __future__ import annotations

import io
from collections import defaultdict
from datetime import date
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.enums import (
    InvoiceStatus,
    PaymentStatus,
    PurchaseInvoiceStatus,
)

if TYPE_CHECKING:
    from app.models import Business, Invoice, Order, PurchaseInvoice


_TITLE = Font(bold=True, size=14, color="0F172A")
_SECTION = Font(bold=True, size=11, color="FFFFFF")
_SECTION_FILL = PatternFill("solid", fgColor="16A34A")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="0F172A")
_RUPEE = '#,##0.00;[Red]-#,##0.00'
_INT = "#,##0"


# Presumptive rates per IT Act
_PRESUMED_CASH_PCT = 8.0
_PRESUMED_DIGITAL_PCT = 6.0
# Eligibility thresholds (FY 2024-25 limits, may update yearly)
_44AD_TURNOVER_LIMIT_PAISE = 2_00_00_000_00   # ₹2 cr
_44ADA_TURNOVER_LIMIT_PAISE = 50_00_000_00    # ₹50 lakh


def _p(paise: int) -> float:
    return round((paise or 0) / 100, 2)


def _section(ws, row: int, label: str, span: int = 6) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = _SECTION
    cell.fill = _SECTION_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22


def _fiscal_quarter(d: date, fy_start_year: int) -> int:
    """Return 1-4 for the Indian FY quarter (Apr-Jun, Jul-Sep, Oct-Dec, Jan-Mar)."""
    # Q1 = Apr-Jun, Q2 = Jul-Sep, Q3 = Oct-Dec, Q4 = Jan-Mar
    if d.year == fy_start_year:
        if d.month >= 4 and d.month <= 6:
            return 1
        if d.month >= 7 and d.month <= 9:
            return 2
        if d.month >= 10 and d.month <= 12:
            return 3
    # Jan-Mar of next year = Q4
    return 4


def _month_label(month: int, year: int) -> str:
    return date(year, month, 1).strftime("%b %Y")


def _classify_payment(order: "Order | None", invoice: "Invoice") -> str:
    """Return 'cash' or 'digital' for presumptive turnover split.

    Heuristic:
      - If linked order has payment_method=cash_on_pickup/delivery → cash
      - Else if order has razorpay_payment_id → digital
      - Else if invoice has razorpay_payment_link → digital
      - Default: cash (safer assumption for presumptive — higher 8% rate)
    """
    if order is not None:
        pm = getattr(order, "payment_method", None)
        pm_val = pm.value if pm and hasattr(pm, "value") else str(pm) if pm else ""
        if "cash" in pm_val:
            return "cash"
        if order.razorpay_payment_id or order.payment_status == PaymentStatus.PAID:
            return "digital"
    if invoice.razorpay_payment_link:
        return "digital"
    return "cash"


# ============================================================
# Sheet builders
# ============================================================


def _sheet_overview(
    ws,
    business: "Business",
    fy_label: str,
    fy_start: date,
    fy_end: date,
    *,
    cash_turnover: int,
    digital_turnover: int,
    expense_total: int,
    inv_count: int,
    is_service: bool,
) -> None:
    """First sheet — high-level numbers + filing guidance."""
    ws.cell(row=1, column=1, value=f"ITR-4 (Sugam) P&L — {business.name}").font = _TITLE
    ws.cell(row=2, column=1, value=f"FY: {fy_label}   ({fy_start.strftime('%d %b %Y')} → {fy_end.strftime('%d %b %Y')})")
    ws.cell(row=3, column=1, value=f"GSTIN: {business.gstin or 'Not registered'}")
    ws.cell(row=4, column=1, value=f"PAN: {business.pan or '—'}")

    gross_turnover = cash_turnover + digital_turnover
    presumed_profit_cash = round(cash_turnover * _PRESUMED_CASH_PCT / 100)
    presumed_profit_digital = round(digital_turnover * _PRESUMED_DIGITAL_PCT / 100)
    presumed_total = presumed_profit_cash + presumed_profit_digital

    r = 6
    _section(ws, r, "Section 44AD / 44ADA — Presumptive Income")
    r += 1
    rows = [
        ("Gross turnover — Cash", cash_turnover, ""),
        ("Gross turnover — Digital (UPI/card/netbanking)", digital_turnover, ""),
        ("Gross turnover — TOTAL", gross_turnover, "Both above combined"),
        ("Presumed profit @8% (cash)", presumed_profit_cash, "Section 44AD"),
        ("Presumed profit @6% (digital)", presumed_profit_digital, "Section 44AD (lower for digital)"),
        ("PRESUMED PROFIT (Total)", presumed_total, "Use this in ITR-4"),
        ("Total expenses recorded", expense_total, "Optional — only if declaring actual profit"),
        ("Actual profit (turnover − expenses)", gross_turnover - expense_total, "Use higher of presumed or actual"),
    ]
    for label, paise, note in rows:
        ws.cell(row=r, column=1, value=label)
        cell = ws.cell(row=r, column=2, value=_p(paise))
        cell.number_format = _RUPEE
        cell.alignment = Alignment(horizontal="right")
        ws.cell(row=r, column=3, value=note).font = Font(italic=True, color="64748b", size=9)
        if "TOTAL" in label or "PRESUMED PROFIT" in label:
            ws.cell(row=r, column=1).font = Font(bold=True)
            ws.cell(row=r, column=2).font = Font(bold=True)
        r += 1

    r += 1
    _section(ws, r, "Eligibility Check")
    r += 1
    limit_paise = _44ADA_TURNOVER_LIMIT_PAISE if is_service else _44AD_TURNOVER_LIMIT_PAISE
    section_name = "44ADA (services)" if is_service else "44AD (goods/business)"
    eligible = gross_turnover <= limit_paise
    ws.cell(row=r, column=1, value=f"Section: {section_name}")
    r += 1
    ws.cell(row=r, column=1, value=f"Turnover limit: ₹{limit_paise / 100:,.0f}")
    r += 1
    ws.cell(
        row=r, column=1,
        value=f"Your turnover: ₹{gross_turnover / 100:,.2f} — {'✅ Eligible for presumptive' if eligible else '❌ Exceeds limit; file ITR-3 instead'}",
    ).font = Font(bold=True, color=("16A34A" if eligible else "DC2626"))
    r += 1

    r += 1
    _section(ws, r, "Invoices Summary")
    r += 1
    ws.cell(row=r, column=1, value="Active invoices (excluding cancelled)")
    ws.cell(row=r, column=2, value=inv_count).number_format = _INT
    r += 1

    r += 1
    ws.cell(
        row=r, column=1,
        value="⚠️ This sheet is a calculation aid. Final ITR-4 filing must be reviewed by a Chartered Accountant. WhatsApp Auto cannot file returns on your behalf.",
    ).font = Font(italic=True, color="DC2626")

    for col, w in enumerate([56, 22, 56], start=1):
        ws.column_dimensions[chr(64 + col)].width = w


def _sheet_quarterly(
    ws,
    invoices: list["Invoice"],
    fy_start_year: int,
) -> None:
    """Q1–Q4 turnover and GST."""
    headers = ["Quarter", "Months", "Invoices", "Taxable (₹)", "Total (₹)", "GST Collected (₹)"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    quarter_buckets: dict[int, dict] = defaultdict(
        lambda: {"count": 0, "taxable": 0, "total": 0, "gst": 0}
    )
    for inv in invoices:
        q = _fiscal_quarter(inv.invoice_date, fy_start_year)
        b = quarter_buckets[q]
        b["count"] += 1
        b["taxable"] += inv.taxable_paise
        b["total"] += inv.total_paise
        b["gst"] += inv.cgst_paise + inv.sgst_paise + inv.igst_paise + inv.cess_paise

    quarter_labels = {1: "Apr–Jun", 2: "Jul–Sep", 3: "Oct–Dec", 4: "Jan–Mar"}
    r = 2
    grand_count = grand_tax = grand_tot = grand_gst = 0
    for q in (1, 2, 3, 4):
        b = quarter_buckets[q]
        ws.cell(row=r, column=1, value=f"Q{q}")
        ws.cell(row=r, column=2, value=quarter_labels[q])
        ws.cell(row=r, column=3, value=b["count"]).number_format = _INT
        for col, key in enumerate(["taxable", "total", "gst"], start=4):
            cell = ws.cell(row=r, column=col, value=_p(b[key]))
            cell.number_format = _RUPEE
            cell.alignment = Alignment(horizontal="right")
        grand_count += b["count"]
        grand_tax += b["taxable"]
        grand_tot += b["total"]
        grand_gst += b["gst"]
        r += 1

    # Total row
    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=r, column=3, value=grand_count).font = Font(bold=True)
    for col, val in zip([4, 5, 6], [grand_tax, grand_tot, grand_gst]):
        cell = ws.cell(row=r, column=col, value=_p(val))
        cell.number_format = _RUPEE
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right")

    ws.freeze_panes = "A2"
    for col, w in enumerate([10, 16, 12, 16, 16, 20], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _sheet_monthly(
    ws,
    invoices: list["Invoice"],
    fy_start_year: int,
) -> None:
    """Sales by month (Apr-Mar order)."""
    headers = ["Month", "Invoices", "Taxable (₹)", "Total (₹)", "GST (₹)"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    month_keys: list[tuple[int, int]] = (
        [(fy_start_year, m) for m in range(4, 13)]
        + [(fy_start_year + 1, m) for m in range(1, 4)]
    )
    buckets: dict[tuple[int, int], dict] = defaultdict(
        lambda: {"count": 0, "taxable": 0, "total": 0, "gst": 0}
    )
    for inv in invoices:
        key = (inv.invoice_date.year, inv.invoice_date.month)
        b = buckets[key]
        b["count"] += 1
        b["taxable"] += inv.taxable_paise
        b["total"] += inv.total_paise
        b["gst"] += inv.cgst_paise + inv.sgst_paise + inv.igst_paise + inv.cess_paise

    r = 2
    for year, month in month_keys:
        b = buckets[(year, month)]
        ws.cell(row=r, column=1, value=_month_label(month, year))
        ws.cell(row=r, column=2, value=b["count"]).number_format = _INT
        for col, key in enumerate(["taxable", "total", "gst"], start=3):
            cell = ws.cell(row=r, column=col, value=_p(b[key]))
            cell.number_format = _RUPEE
            cell.alignment = Alignment(horizontal="right")
        r += 1

    ws.freeze_panes = "A2"
    for col, w in enumerate([14, 12, 16, 16, 16], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _sheet_expenses(
    ws,
    purchase_invoices: list["PurchaseInvoice"],
) -> None:
    """Purchase totals by month + category breakdown."""
    headers = ["Bill Date", "Supplier", "Category", "Capital?", "Taxable (₹)", "Total (₹)"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    r = 2
    total_paise = 0
    capital_paise = 0
    by_category: dict[str, int] = defaultdict(int)

    for p in sorted(purchase_invoices, key=lambda x: x.bill_date):
        ws.cell(row=r, column=1, value=p.bill_date).number_format = "dd-mmm-yyyy"
        ws.cell(row=r, column=2, value=p.supplier_name)
        ws.cell(row=r, column=3, value=p.category or "uncategorized")
        ws.cell(row=r, column=4, value="Yes" if p.is_capital_goods else "No")
        ws.cell(row=r, column=5, value=_p(p.taxable_paise)).number_format = _RUPEE
        ws.cell(row=r, column=6, value=_p(p.total_paise)).number_format = _RUPEE
        total_paise += p.total_paise
        if p.is_capital_goods:
            capital_paise += p.total_paise
        by_category[p.category or "uncategorized"] += p.total_paise
        r += 1

    if r > 2:
        ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
        cell = ws.cell(row=r, column=6, value=_p(total_paise))
        cell.font = Font(bold=True)
        cell.number_format = _RUPEE
        r += 2

    # Category summary
    ws.cell(row=r, column=1, value="Category Breakdown").font = Font(bold=True, size=11)
    r += 1
    ws.cell(row=r, column=1, value="Category").font = Font(bold=True)
    ws.cell(row=r, column=2, value="Total (₹)").font = Font(bold=True)
    r += 1
    for cat, total in sorted(by_category.items(), key=lambda x: -x[1]):
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2, value=_p(total)).number_format = _RUPEE
        r += 1
    if capital_paise:
        r += 1
        ws.cell(row=r, column=1, value=f"Of which capital goods: ₹{capital_paise / 100:,.2f}").font = Font(italic=True)

    ws.freeze_panes = "A2"
    for col, w in enumerate([14, 28, 18, 10, 16, 16], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _sheet_filing_worksheet(
    ws,
    fy_label: str,
    *,
    gross_turnover: int,
    presumed_profit: int,
    capital_total: int,
) -> None:
    """Maps our numbers to ITR-4 schedule line names."""
    ws.cell(row=1, column=1, value=f"ITR-4 Filing Worksheet — FY {fy_label}").font = _TITLE
    ws.cell(
        row=2, column=1,
        value="Copy these numbers into the corresponding ITR-4 fields on incometax.gov.in",
    ).font = Font(italic=True, color="64748b")

    rows = [
        ("PART A — Personal Info", ""),
        ("    Filing under: Section 44AD (or 44ADA for services)", ""),
        ("", ""),
        ("PART B — Gross Total Income", ""),
        ("    Schedule BP — Income from Business (presumptive)", ""),
        ("    Gross turnover/receipts (E1+E2)", _p(gross_turnover)),
        ("        Cash receipts (E1)", "computed in Overview"),
        ("        Other than cash (E2)", "computed in Overview"),
        ("    Presumed profit (E11)", _p(presumed_profit)),
        ("", ""),
        ("PART C — Tax Computation", ""),
        ("    Net taxable income", _p(presumed_profit)),
        ("    Add: other income (interest, etc.)", "manual"),
        ("    Less: Deductions u/s 80C/80D/etc.", "manual"),
        ("", ""),
        ("Schedule AL — Asset & Liabilities (mandatory if total income > ₹50L)", ""),
        ("    Capital goods recorded this FY", _p(capital_total)),
    ]
    r = 4
    for label, val in rows:
        ws.cell(row=r, column=1, value=label)
        if label.strip().startswith(("PART", "Schedule")):
            ws.cell(row=r, column=1).font = Font(bold=True)
        if val != "":
            cell = ws.cell(row=r, column=2, value=val)
            if isinstance(val, float):
                cell.number_format = _RUPEE
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.font = Font(italic=True, color="64748b")
        r += 1

    for col, w in enumerate([56, 22], start=1):
        ws.column_dimensions[chr(64 + col)].width = w


# ============================================================
# Public entrypoint
# ============================================================


def build_itr4_pnl(
    business: "Business",
    invoices: list["Invoice"],
    purchase_invoices: list["PurchaseInvoice"],
    fy_label: str,
    fy_start: date,
    fy_end: date,
    *,
    orders_by_invoice: dict | None = None,
) -> bytes:
    """Build the ITR-4 P&L xlsx.

    Args:
        invoices: ALL invoices in the FY (we filter cancelled internally)
        purchase_invoices: ALL purchase bills in the FY
        fy_label: e.g. "2026-27"
        fy_start: date(2026, 4, 1)
        fy_end:   date(2027, 3, 31)
        orders_by_invoice: optional map of {invoice_id: Order} to classify
                           cash vs digital based on order.payment_method
    """
    active_inv = [i for i in invoices if i.status != InvoiceStatus.CANCELLED]
    active_purchases = [
        p for p in purchase_invoices if p.status != PurchaseInvoiceStatus.CANCELLED
    ]
    orders_by_invoice = orders_by_invoice or {}

    # Split turnover by payment method
    cash_paise = 0
    digital_paise = 0
    for inv in active_inv:
        order = orders_by_invoice.get(inv.id) if hasattr(inv, "id") else None
        if _classify_payment(order, inv) == "cash":
            cash_paise += inv.total_paise
        else:
            digital_paise += inv.total_paise

    expense_total = sum(p.total_paise for p in active_purchases)
    capital_total = sum(p.total_paise for p in active_purchases if p.is_capital_goods)
    gross_turnover = cash_paise + digital_paise
    presumed = round(cash_paise * _PRESUMED_CASH_PCT / 100) + round(
        digital_paise * _PRESUMED_DIGITAL_PCT / 100
    )

    # Detect if predominantly a service biz (44ADA) — heuristic: business_type
    biz_type = getattr(business, "business_type", None)
    bt_val = (
        biz_type.value if biz_type and hasattr(biz_type, "value") else str(biz_type or "")
    )
    is_service = bt_val in ("salon", "clinic", "coaching", "agency")

    wb = Workbook()
    _sheet_overview(
        wb.active, business, fy_label, fy_start, fy_end,
        cash_turnover=cash_paise, digital_turnover=digital_paise,
        expense_total=expense_total, inv_count=len(active_inv),
        is_service=is_service,
    )
    wb.active.title = "Overview"
    _sheet_quarterly(wb.create_sheet("Quarterly"), active_inv, fy_start.year)
    _sheet_monthly(wb.create_sheet("Monthly Sales"), active_inv, fy_start.year)
    _sheet_expenses(wb.create_sheet("Expenses"), active_purchases)
    _sheet_filing_worksheet(
        wb.create_sheet("Filing Worksheet"),
        fy_label,
        gross_turnover=gross_turnover,
        presumed_profit=presumed,
        capital_total=capital_total,
    )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
