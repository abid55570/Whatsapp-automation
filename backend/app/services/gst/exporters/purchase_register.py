"""Purchase register xlsx — inward supplies for ITC tracking + GSTR-2A reconciliation."""
from __future__ import annotations

import io
from datetime import date
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.models.enums import PurchaseInvoiceStatus

if TYPE_CHECKING:
    from app.models import Business, PurchaseInvoice


_HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="16A34A")
_RUPEE_FORMAT = '#,##0.00;[Red]-#,##0.00'


def _p(paise: int) -> float:
    return round((paise or 0) / 100, 2)


def build_purchase_register(
    business: "Business",
    purchase_invoices: list["PurchaseInvoice"],
    period_from: date,
    period_to: date,
) -> bytes:
    active = [p for p in purchase_invoices if p.status != PurchaseInvoiceStatus.CANCELLED]

    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Register"

    ws.cell(row=1, column=1, value=f"Purchase Register — {business.name}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"GSTIN: {business.gstin or 'Not registered'}")
    ws.cell(row=3, column=1, value=f"Period: {period_from.strftime('%d %b %Y')} → {period_to.strftime('%d %b %Y')}")
    ws.cell(row=4, column=1, value=f"Bills: {len(active)} active · {len(purchase_invoices) - len(active)} cancelled")

    headers = [
        "Date", "Bill No.", "Supplier", "Supplier GSTIN", "State",
        "Category", "Capital?", "ITC?",
        "Taxable (₹)", "IGST (₹)", "CGST (₹)", "SGST (₹)", "Cess (₹)", "Total (₹)",
    ]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=c, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    r = 7
    totals = {"taxable": 0, "igst": 0, "cgst": 0, "sgst": 0, "cess": 0, "total": 0}
    for p in active:
        row = [
            p.bill_date,
            p.bill_number,
            p.supplier_name,
            p.supplier_gstin or "",
            p.supplier_state_code or "",
            p.category or "",
            "Yes" if p.is_capital_goods else "No",
            "Yes" if p.is_itc_eligible else "No",
            _p(p.taxable_paise),
            _p(p.igst_paise),
            _p(p.cgst_paise),
            _p(p.sgst_paise),
            _p(p.cess_paise),
            _p(p.total_paise),
        ]
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            if c == 1:
                cell.number_format = "dd-mmm-yyyy"
            elif c >= 9:
                cell.number_format = _RUPEE_FORMAT
                cell.alignment = Alignment(horizontal="right")
        totals["taxable"] += p.taxable_paise
        totals["igst"] += p.igst_paise
        totals["cgst"] += p.cgst_paise
        totals["sgst"] += p.sgst_paise
        totals["cess"] += p.cess_paise
        totals["total"] += p.total_paise
        r += 1

    # Total row
    if r > 7:
        ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
        for c, key in enumerate(
            ["taxable", "igst", "cgst", "sgst", "cess", "total"], start=9
        ):
            cell = ws.cell(row=r, column=c, value=_p(totals[key]))
            cell.number_format = _RUPEE_FORMAT
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="right")

    ws.freeze_panes = "A7"
    for col, w in enumerate(
        [12, 14, 28, 18, 8, 14, 10, 8, 14, 14, 14, 14, 12, 14], start=1
    ):
        ws.column_dimensions[chr(64 + col)].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
