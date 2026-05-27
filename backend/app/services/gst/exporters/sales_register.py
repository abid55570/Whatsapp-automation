"""Monthly sales register — multi-sheet xlsx for the owner / their CA.

Output sheets:
    1. Sales Register   — every invoice line, flat
    2. Invoice Summary  — one row per invoice
    3. Tax Summary      — totals by GST rate
    4. B2B              — only B2B invoices (cx_gstin present)
    5. B2C Large        — interstate B2C > ₹2.5L
    6. HSN Summary      — HSN-wise totals (for GSTR-1)

Built with openpyxl directly (no pandas → smaller install, more control over formatting).
"""
from __future__ import annotations

import io
import logging
from collections import defaultdict
from datetime import date
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.enums import InvoiceStatus, InvoiceType

if TYPE_CHECKING:
    from app.models import Business, Invoice

logger = logging.getLogger(__name__)


# ============================================================
# Styling helpers
# ============================================================


_HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="16A34A")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_RUPEE_FORMAT = '#,##0.00;[Red]-#,##0.00'
_INT_FORMAT = '#,##0'
_DATE_FORMAT = "dd-mmm-yyyy"


def _style_header(cell):
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.alignment = _HEADER_ALIGN


def _autofit(ws, max_width: int = 50):
    """Best-effort column auto-width."""
    for col in ws.columns:
        length = 8
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = str(cell.value) if cell.value is not None else ""
                length = max(length, min(max_width, len(v) + 2))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = length


def _paise_to_inr(p: int) -> float:
    return round((p or 0) / 100, 2)


# ============================================================
# Sheet builders
# ============================================================


def _sheet_invoice_summary(ws, invoices: list["Invoice"]) -> None:
    headers = [
        "Invoice No.", "Date", "FY", "Customer", "GSTIN", "Place of Supply",
        "Type", "Status", "Taxable (₹)", "CGST (₹)", "SGST (₹)", "IGST (₹)",
        "Cess (₹)", "Round Off (₹)", "Total (₹)",
    ]
    for col, h in enumerate(headers, start=1):
        _style_header(ws.cell(row=1, column=col, value=h))

    for r, inv in enumerate(invoices, start=2):
        row = [
            inv.invoice_number,
            inv.invoice_date,
            inv.fiscal_year,
            inv.cx_name or "",
            inv.cx_gstin or "",
            inv.place_of_supply or "",
            inv.invoice_type.value if hasattr(inv.invoice_type, "value") else str(inv.invoice_type),
            inv.status.value if hasattr(inv.status, "value") else str(inv.status),
            _paise_to_inr(inv.taxable_paise),
            _paise_to_inr(inv.cgst_paise),
            _paise_to_inr(inv.sgst_paise),
            _paise_to_inr(inv.igst_paise),
            _paise_to_inr(inv.cess_paise),
            _paise_to_inr(inv.round_off_paise),
            _paise_to_inr(inv.total_paise),
        ]
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            if c == 2:  # date
                cell.number_format = _DATE_FORMAT
            elif c >= 9:  # money columns
                cell.number_format = _RUPEE_FORMAT
                cell.alignment = Alignment(horizontal="right")

    ws.freeze_panes = "A2"
    _autofit(ws)


def _sheet_line_register(ws, invoices: list["Invoice"]) -> None:
    headers = [
        "Invoice No.", "Date", "Line", "Description", "HSN", "Qty", "Unit",
        "Rate (₹)", "Disc %", "GST %", "Taxable (₹)", "CGST (₹)", "SGST (₹)",
        "IGST (₹)", "Cess (₹)", "Total (₹)",
    ]
    for col, h in enumerate(headers, start=1):
        _style_header(ws.cell(row=1, column=col, value=h))

    r = 2
    for inv in invoices:
        for line in inv.lines:
            row = [
                inv.invoice_number,
                inv.invoice_date,
                line.line_number,
                line.description,
                line.hsn_code or "",
                float(line.quantity),
                line.unit,
                _paise_to_inr(line.rate_paise),
                float(line.discount_pct),
                line.gst_rate,
                _paise_to_inr(line.taxable_paise),
                _paise_to_inr(line.cgst_paise),
                _paise_to_inr(line.sgst_paise),
                _paise_to_inr(line.igst_paise),
                _paise_to_inr(line.cess_paise),
                _paise_to_inr(line.total_paise),
            ]
            for c, v in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                if c == 2:
                    cell.number_format = _DATE_FORMAT
                elif c in (8, 11, 12, 13, 14, 15, 16):
                    cell.number_format = _RUPEE_FORMAT
                    cell.alignment = Alignment(horizontal="right")
            r += 1
    ws.freeze_panes = "A2"
    _autofit(ws)


def _sheet_tax_summary(ws, invoices: list["Invoice"]) -> None:
    """Totals grouped by GST rate."""
    headers = ["GST Rate", "Invoices", "Taxable (₹)", "CGST (₹)", "SGST (₹)", "IGST (₹)", "Total Tax (₹)"]
    for col, h in enumerate(headers, start=1):
        _style_header(ws.cell(row=1, column=col, value=h))

    # rate → [count, taxable, cgst, sgst, igst]
    buckets: dict[int, list[int]] = defaultdict(lambda: [0, 0, 0, 0, 0])
    counted_invoices: dict[int, set] = defaultdict(set)
    for inv in invoices:
        for line in inv.lines:
            b = buckets[line.gst_rate]
            counted_invoices[line.gst_rate].add(inv.id)
            b[1] += line.taxable_paise
            b[2] += line.cgst_paise
            b[3] += line.sgst_paise
            b[4] += line.igst_paise

    r = 2
    grand = [0, 0, 0, 0, 0]
    for rate in sorted(buckets.keys()):
        b = buckets[rate]
        n = len(counted_invoices[rate])
        ws.cell(row=r, column=1, value=f"{rate}%")
        ws.cell(row=r, column=2, value=n).number_format = _INT_FORMAT
        ws.cell(row=r, column=3, value=_paise_to_inr(b[1])).number_format = _RUPEE_FORMAT
        ws.cell(row=r, column=4, value=_paise_to_inr(b[2])).number_format = _RUPEE_FORMAT
        ws.cell(row=r, column=5, value=_paise_to_inr(b[3])).number_format = _RUPEE_FORMAT
        ws.cell(row=r, column=6, value=_paise_to_inr(b[4])).number_format = _RUPEE_FORMAT
        tax = b[2] + b[3] + b[4]
        ws.cell(row=r, column=7, value=_paise_to_inr(tax)).number_format = _RUPEE_FORMAT
        for i in range(5):
            grand[i] += b[i] if i == 0 else b[i]
        r += 1

    # Grand total row
    if r > 2:
        ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=r, column=2, value=sum(len(s) for s in counted_invoices.values())).font = Font(bold=True)
        for c, val_paise in enumerate(grand[1:], start=3):
            cell = ws.cell(row=r, column=c, value=_paise_to_inr(val_paise))
            cell.number_format = _RUPEE_FORMAT
            cell.font = Font(bold=True)
        tot_tax = grand[2] + grand[3] + grand[4]
        ws.cell(row=r, column=7, value=_paise_to_inr(tot_tax)).number_format = _RUPEE_FORMAT
    ws.freeze_panes = "A2"
    _autofit(ws)


def _sheet_b2b(ws, invoices: list["Invoice"]) -> None:
    b2b = [i for i in invoices if i.invoice_type == InvoiceType.B2B]
    _sheet_invoice_summary(ws, b2b)


def _sheet_b2c_large(ws, invoices: list["Invoice"]) -> None:
    b2cl = [i for i in invoices if i.invoice_type == InvoiceType.B2C_LARGE]
    _sheet_invoice_summary(ws, b2cl)


def _sheet_hsn_summary(ws, invoices: list["Invoice"]) -> None:
    """HSN-wise totals — required section in GSTR-1."""
    headers = [
        "HSN", "Description (sample)", "UQC (unit)", "Total Qty",
        "Taxable (₹)", "IGST (₹)", "CGST (₹)", "SGST (₹)", "Cess (₹)",
    ]
    for col, h in enumerate(headers, start=1):
        _style_header(ws.cell(row=1, column=col, value=h))

    # hsn → bucket
    buckets: dict[str, dict] = {}
    for inv in invoices:
        for line in inv.lines:
            hsn = line.hsn_code or "UNCLASSIFIED"
            b = buckets.setdefault(
                hsn,
                {"desc": line.description, "unit": line.unit, "qty": 0.0,
                 "taxable": 0, "igst": 0, "cgst": 0, "sgst": 0, "cess": 0},
            )
            b["qty"] += float(line.quantity)
            b["taxable"] += line.taxable_paise
            b["igst"] += line.igst_paise
            b["cgst"] += line.cgst_paise
            b["sgst"] += line.sgst_paise
            b["cess"] += line.cess_paise

    r = 2
    for hsn in sorted(buckets.keys()):
        b = buckets[hsn]
        ws.cell(row=r, column=1, value=hsn)
        ws.cell(row=r, column=2, value=b["desc"])
        ws.cell(row=r, column=3, value=b["unit"])
        ws.cell(row=r, column=4, value=round(b["qty"], 3))
        ws.cell(row=r, column=5, value=_paise_to_inr(b["taxable"])).number_format = _RUPEE_FORMAT
        ws.cell(row=r, column=6, value=_paise_to_inr(b["igst"])).number_format = _RUPEE_FORMAT
        ws.cell(row=r, column=7, value=_paise_to_inr(b["cgst"])).number_format = _RUPEE_FORMAT
        ws.cell(row=r, column=8, value=_paise_to_inr(b["sgst"])).number_format = _RUPEE_FORMAT
        ws.cell(row=r, column=9, value=_paise_to_inr(b["cess"])).number_format = _RUPEE_FORMAT
        r += 1
    ws.freeze_panes = "A2"
    _autofit(ws)


# ============================================================
# Public entrypoint
# ============================================================


def build_sales_register(
    business: "Business",
    invoices: list["Invoice"],
    period_from: date,
    period_to: date,
) -> bytes:
    """Build the full multi-sheet xlsx. Returns the binary payload.

    `invoices` must be pre-filtered to the desired date range AND eager-loaded
    with `.lines` (we read them sync here).
    """
    # Exclude cancelled invoices from filing — they don't enter GSTR
    active = [i for i in invoices if i.status != InvoiceStatus.CANCELLED]

    wb = Workbook()
    # Cover sheet
    cover = wb.active
    cover.title = "Overview"
    cover.cell(row=1, column=1, value=f"Sales Register — {business.name}").font = Font(bold=True, size=14)
    cover.cell(row=2, column=1, value=f"GSTIN: {business.gstin or 'Unregistered'}")
    cover.cell(row=3, column=1, value=f"Period: {period_from.strftime('%d %b %Y')} → {period_to.strftime('%d %b %Y')}")
    cover.cell(row=4, column=1, value=f"Invoices: {len(active)} active · {len(invoices) - len(active)} cancelled (excluded)")
    cover.cell(row=5, column=1, value="Generated by WhatsApp Auto")
    cover.cell(row=6, column=1, value="").font = Font(italic=True, color="64748b")
    cover.cell(
        row=7, column=1,
        value="NOTE: Verify all figures with your CA before filing. WhatsApp Auto is not a substitute for professional tax advice.",
    ).font = Font(italic=True, color="DC2626")

    _sheet_invoice_summary(wb.create_sheet("Invoice Summary"), active)
    _sheet_line_register(wb.create_sheet("Line Register"), active)
    _sheet_tax_summary(wb.create_sheet("Tax Summary"), active)
    _sheet_b2b(wb.create_sheet("B2B"), active)
    _sheet_b2c_large(wb.create_sheet("B2C Large"), active)
    _sheet_hsn_summary(wb.create_sheet("HSN Summary"), active)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
