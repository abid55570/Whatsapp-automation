"""Tests for the ITR-4 P&L exporter."""
from datetime import date
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from app.models.enums import (
    BusinessType,
    GstScheme,
    InvoiceStatus,
    InvoiceType,
    PaymentMethod,
    PaymentStatus,
    PurchaseInvoiceStatus,
)
from app.services.gst.exporters import itr4_pnl


def _line(**kw):
    base = dict(
        line_number=1, description="Item", hsn_code="1001",
        quantity=Decimal("1"), unit="pc", rate_paise=10000,
        discount_pct=Decimal("0"), gst_rate=18,
        taxable_paise=10000, cgst_paise=900, sgst_paise=900,
        igst_paise=0, cess_paise=0, total_paise=11800,
    )
    base.update(kw)
    return SimpleNamespace(**base)


def _inv(n=1, **kw):
    base = dict(
        id=f"id-{n}",
        invoice_number=f"INV-26-{n:04d}",
        invoice_date=date(2026, 5, 15),
        fiscal_year="2026-27",
        cx_state_code="27",
        subtotal_paise=10000, discount_paise=0,
        taxable_paise=10000, cgst_paise=900, sgst_paise=900,
        igst_paise=0, cess_paise=0, round_off_paise=0, total_paise=11800,
        razorpay_payment_link=None,
        invoice_type=InvoiceType.B2C,
        status=InvoiceStatus.ISSUED,
        order_id=None,
        lines=[_line()],
    )
    base.update(kw)
    return SimpleNamespace(**base)


def _purchase(**kw):
    base = dict(
        supplier_name="S", bill_number="B1", bill_date=date(2026, 6, 1),
        category="raw_materials", is_capital_goods=False,
        is_itc_eligible=True, status=PurchaseInvoiceStatus.RECORDED,
        taxable_paise=5000, cgst_paise=450, sgst_paise=450,
        igst_paise=0, cess_paise=0, total_paise=5900,
    )
    base.update(kw)
    return SimpleNamespace(**base)


def _biz(**kw):
    base = dict(
        name="Sharma Kirana",
        gstin="27AAPFU0939F1ZV",
        pan="AAPFU0939F",
        gst_state_code="27",
        gst_scheme=GstScheme.REGULAR,
        business_type=BusinessType.SHOP,
    )
    base.update(kw)
    return SimpleNamespace(**base)


class TestITR4PnL:
    def test_produces_valid_xlsx(self):
        out = itr4_pnl.build_itr4_pnl(
            _biz(), [_inv()], [_purchase()],
            "2026-27", date(2026, 4, 1), date(2027, 3, 31),
        )
        assert out[:2] == b"PK"
        wb = load_workbook(BytesIO(out))
        assert "Overview" in wb.sheetnames
        assert "Quarterly" in wb.sheetnames
        assert "Monthly Sales" in wb.sheetnames
        assert "Expenses" in wb.sheetnames
        assert "Filing Worksheet" in wb.sheetnames

    def test_overview_shows_turnover(self):
        out = itr4_pnl.build_itr4_pnl(
            _biz(), [_inv()], [],
            "2026-27", date(2026, 4, 1), date(2027, 3, 31),
        )
        wb = load_workbook(BytesIO(out))
        ws = wb["Overview"]
        # 11800 paise = ₹118.00 — should appear somewhere
        all_nums: set[float] = set()
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    all_nums.add(round(float(v), 2))
        assert 118.0 in all_nums

    def test_excludes_cancelled(self):
        active = _inv(1)
        cancelled = _inv(2, status=InvoiceStatus.CANCELLED, total_paise=99999)
        out = itr4_pnl.build_itr4_pnl(
            _biz(), [active, cancelled], [],
            "2026-27", date(2026, 4, 1), date(2027, 3, 31),
        )
        wb = load_workbook(BytesIO(out))
        ws = wb["Overview"]
        nums: set[float] = set()
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    nums.add(round(float(v), 2))
        # Cancelled value should not appear
        assert 999.99 not in nums

    def test_quarterly_split(self):
        invs = [
            _inv(1, invoice_date=date(2026, 5, 1)),     # Q1
            _inv(2, invoice_date=date(2026, 8, 1)),     # Q2
            _inv(3, invoice_date=date(2026, 11, 1)),    # Q3
            _inv(4, invoice_date=date(2027, 2, 1)),     # Q4
        ]
        out = itr4_pnl.build_itr4_pnl(
            _biz(), invs, [],
            "2026-27", date(2026, 4, 1), date(2027, 3, 31),
        )
        wb = load_workbook(BytesIO(out))
        ws = wb["Quarterly"]
        # Each Q row should show 1 invoice (column 3)
        for r in range(2, 6):
            assert ws.cell(row=r, column=3).value == 1

    def test_service_business_uses_44ada(self):
        # Salon → service biz → 44ADA threshold
        out = itr4_pnl.build_itr4_pnl(
            _biz(business_type=BusinessType.SALON), [_inv()], [],
            "2026-27", date(2026, 4, 1), date(2027, 3, 31),
        )
        wb = load_workbook(BytesIO(out))
        ws = wb["Overview"]
        found_44ada = False
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str) and "44ADA" in v:
                    found_44ada = True
                    break
        assert found_44ada

    def test_expenses_sheet_lists_purchases(self):
        purchases = [
            _purchase(supplier_name="A", category="raw"),
            _purchase(supplier_name="B", category="office",
                      is_capital_goods=True, total_paise=50000),
        ]
        out = itr4_pnl.build_itr4_pnl(
            _biz(), [_inv()], purchases,
            "2026-27", date(2026, 4, 1), date(2027, 3, 31),
        )
        wb = load_workbook(BytesIO(out))
        ws = wb["Expenses"]
        # Row 2 = first supplier, row 3 = second
        assert ws.cell(row=2, column=2).value == "A"
        assert ws.cell(row=3, column=2).value == "B"
        assert ws.cell(row=3, column=4).value == "Yes"   # capital goods


class TestPaymentClassification:
    def test_cash_order(self):
        order = SimpleNamespace(
            payment_method=PaymentMethod.CASH_ON_PICKUP,
            payment_status=PaymentStatus.PAID,
            razorpay_payment_id=None,
        )
        assert itr4_pnl._classify_payment(order, _inv()) == "cash"

    def test_digital_order(self):
        order = SimpleNamespace(
            payment_method=PaymentMethod.UPI,
            payment_status=PaymentStatus.PAID,
            razorpay_payment_id="pay_xyz",
        )
        assert itr4_pnl._classify_payment(order, _inv()) == "digital"

    def test_invoice_with_razorpay_link_digital(self):
        inv = _inv(razorpay_payment_link="https://rzp.io/i/abc")
        assert itr4_pnl._classify_payment(None, inv) == "digital"

    def test_default_to_cash(self):
        # No order, no payment link
        assert itr4_pnl._classify_payment(None, _inv()) == "cash"
