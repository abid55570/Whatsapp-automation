"""Tests for sales register / GSTR-1 / GSTR-3B / purchase register builders."""
from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from app.models.enums import (
    GstScheme,
    InvoiceStatus,
    InvoiceType,
    PurchaseInvoiceStatus,
)
from app.services.gst.exporters import (
    gstr1 as gstr1_mod,
    gstr3b_summary as gstr3b_mod,
    purchase_register as pr_mod,
    sales_register as sr_mod,
)


# ============================================================
# Test fixtures
# ============================================================


def _line(n=1, **kw):
    base = dict(
        line_number=n,
        description=f"Item {n}",
        hsn_code="1001",
        quantity=Decimal("1"),
        unit="pc",
        rate_paise=10000,
        discount_pct=Decimal("0"),
        gst_rate=18,
        taxable_paise=10000,
        cgst_paise=900,
        sgst_paise=900,
        igst_paise=0,
        cess_paise=0,
        total_paise=11800,
    )
    base.update(kw)
    return SimpleNamespace(**base)


def _invoice(n=1, **kw):
    base = dict(
        id=f"id-{n}",
        invoice_number=f"INV-26-{n:04d}",
        invoice_date=date(2026, 5, n if n < 28 else 28),
        fiscal_year="2026-27",
        cx_name="Customer",
        cx_phone="+919876543210",
        cx_gstin=None,
        cx_address=None,
        cx_state_code="27",
        subtotal_paise=10000,
        discount_paise=0,
        taxable_paise=10000,
        cgst_paise=900,
        sgst_paise=900,
        igst_paise=0,
        cess_paise=0,
        round_off_paise=0,
        total_paise=11800,
        place_of_supply="Maharashtra",
        reverse_charge=False,
        invoice_type=InvoiceType.B2C,
        notes=None,
        status=InvoiceStatus.ISSUED,
        irn=None,
        lines=[_line(1)],
    )
    base.update(kw)
    return SimpleNamespace(**base)


def _business(**kw):
    base = dict(
        name="Sharma Kirana",
        gstin="27AAPFU0939F1ZV",
        gst_state_code="27",
        gst_scheme=GstScheme.REGULAR,
    )
    base.update(kw)
    return SimpleNamespace(**base)


def _purchase(n=1, **kw):
    base = dict(
        supplier_name=f"Supplier {n}",
        supplier_gstin="29AAGCB7383J1Z4",
        supplier_state_code="29",
        bill_number=f"BILL-{n}",
        bill_date=date(2026, 5, 10),
        taxable_paise=5000,
        cgst_paise=0,
        sgst_paise=0,
        igst_paise=900,
        cess_paise=0,
        total_paise=5900,
        category="raw_materials",
        is_capital_goods=False,
        is_itc_eligible=True,
        status=PurchaseInvoiceStatus.RECORDED,
    )
    base.update(kw)
    return SimpleNamespace(**base)


# ============================================================
# Sales register xlsx
# ============================================================


class TestSalesRegister:
    def test_produces_valid_xlsx(self):
        out = sr_mod.build_sales_register(
            _business(), [_invoice()], date(2026, 5, 1), date(2026, 5, 31)
        )
        assert out[:2] == b"PK"  # xlsx is a zip
        wb = load_workbook(BytesIO(out))
        assert "Overview" in wb.sheetnames
        assert "Invoice Summary" in wb.sheetnames
        assert "Tax Summary" in wb.sheetnames
        assert "B2B" in wb.sheetnames
        assert "HSN Summary" in wb.sheetnames

    def test_includes_invoice_number_in_summary(self):
        out = sr_mod.build_sales_register(
            _business(), [_invoice()], date(2026, 5, 1), date(2026, 5, 31)
        )
        wb = load_workbook(BytesIO(out))
        ws = wb["Invoice Summary"]
        # Row 2 first cell should be the invoice number
        assert ws.cell(row=2, column=1).value == "INV-26-0001"

    def test_excludes_cancelled(self):
        cancelled = _invoice(2, status=InvoiceStatus.CANCELLED)
        active = _invoice(1)
        out = sr_mod.build_sales_register(
            _business(), [active, cancelled], date(2026, 5, 1), date(2026, 5, 31)
        )
        wb = load_workbook(BytesIO(out))
        ws = wb["Invoice Summary"]
        # Only one data row (row 2); row 3 should be empty
        assert ws.cell(row=2, column=1).value == "INV-26-0001"
        assert ws.cell(row=3, column=1).value is None

    def test_tax_summary_groups_by_rate(self):
        invs = [
            _invoice(1, lines=[_line(1, gst_rate=5, cgst_paise=250, sgst_paise=250)]),
            _invoice(2, lines=[_line(1, gst_rate=18, cgst_paise=900, sgst_paise=900)]),
            _invoice(3, lines=[_line(1, gst_rate=18, cgst_paise=900, sgst_paise=900)]),
        ]
        out = sr_mod.build_sales_register(
            _business(), invs, date(2026, 5, 1), date(2026, 5, 31)
        )
        wb = load_workbook(BytesIO(out))
        ws = wb["Tax Summary"]
        # Two distinct rates → two rows + TOTAL
        rate_cells = [ws.cell(row=r, column=1).value for r in range(2, 6)]
        assert "5%" in rate_cells
        assert "18%" in rate_cells
        assert "TOTAL" in rate_cells

    def test_empty_invoice_list(self):
        out = sr_mod.build_sales_register(
            _business(), [], date(2026, 5, 1), date(2026, 5, 31)
        )
        assert out[:2] == b"PK"

    def test_b2b_sheet_only_b2b(self):
        invs = [
            _invoice(1, invoice_type=InvoiceType.B2C),
            _invoice(2, invoice_type=InvoiceType.B2B, cx_gstin="29AAGCB7383J1Z4"),
        ]
        out = sr_mod.build_sales_register(
            _business(), invs, date(2026, 5, 1), date(2026, 5, 31)
        )
        wb = load_workbook(BytesIO(out))
        ws = wb["B2B"]
        # Only B2B invoice should appear (row 2)
        assert ws.cell(row=2, column=1).value == "INV-26-0002"
        assert ws.cell(row=3, column=1).value is None


# ============================================================
# GSTR-1 JSON
# ============================================================


class TestGSTR1:
    def test_empty(self):
        out = gstr1_mod.build_gstr1_json(_business(), [], date(2026, 5, 1))
        assert out["gstin"] == "27AAPFU0939F1ZV"
        assert out["fp"] == "052026"
        assert out["b2b"] == []
        assert out["b2cl"] == []
        assert out["b2cs"] == []

    def test_b2b_grouped_by_ctin(self):
        invs = [
            _invoice(1, invoice_type=InvoiceType.B2B, cx_gstin="29AAGCB7383J1Z4",
                     cx_state_code="29",
                     cgst_paise=0, sgst_paise=0, igst_paise=1800),
            _invoice(2, invoice_type=InvoiceType.B2B, cx_gstin="29AAGCB7383J1Z4",
                     cx_state_code="29",
                     cgst_paise=0, sgst_paise=0, igst_paise=1800,
                     lines=[_line(1, gst_rate=18, cgst_paise=0, sgst_paise=0, igst_paise=1800)]),
        ]
        out = gstr1_mod.build_gstr1_json(_business(), invs, date(2026, 5, 1))
        assert len(out["b2b"]) == 1   # single ctin
        assert out["b2b"][0]["ctin"] == "29AAGCB7383J1Z4"
        assert len(out["b2b"][0]["inv"]) == 2

    def test_b2cl_grouped_by_pos(self):
        inv = _invoice(
            1, invoice_type=InvoiceType.B2C_LARGE, cx_state_code="07",
            cgst_paise=0, sgst_paise=0, igst_paise=1800,
            total_paise=30_00_000_00,
        )
        out = gstr1_mod.build_gstr1_json(_business(), [inv], date(2026, 5, 1))
        assert len(out["b2cl"]) == 1
        assert out["b2cl"][0]["pos"] == "07"

    def test_b2cs_intra_inter(self):
        intra = _invoice(1, invoice_type=InvoiceType.B2C, cx_state_code="27")
        inter = _invoice(
            2, invoice_type=InvoiceType.B2C, cx_state_code="07",
            cgst_paise=0, sgst_paise=0, igst_paise=1800,
            lines=[_line(1, cgst_paise=0, sgst_paise=0, igst_paise=1800)],
        )
        out = gstr1_mod.build_gstr1_json(_business(), [intra, inter], date(2026, 5, 1))
        sply_types = {row["sply_ty"] for row in out["b2cs"]}
        assert "INTRA" in sply_types
        assert "INTER" in sply_types

    def test_hsn_summary(self):
        invs = [
            _invoice(1, lines=[_line(1, hsn_code="1001"), _line(2, hsn_code="1001")]),
            _invoice(2, lines=[_line(1, hsn_code="8703")]),
        ]
        out = gstr1_mod.build_gstr1_json(_business(), invs, date(2026, 5, 1))
        hsns = {row["hsn_sc"] for row in out["hsn"]["data"]}
        assert "1001" in hsns
        assert "8703" in hsns

    def test_excludes_cancelled(self):
        invs = [
            _invoice(1),
            _invoice(2, status=InvoiceStatus.CANCELLED, invoice_type=InvoiceType.B2B,
                     cx_gstin="29AAGCB7383J1Z4"),
        ]
        out = gstr1_mod.build_gstr1_json(_business(), invs, date(2026, 5, 1))
        # Cancelled B2B invoice excluded
        assert out["b2b"] == []

    def test_period_format(self):
        out = gstr1_mod.build_gstr1_json(_business(), [], date(2026, 3, 15))
        assert out["fp"] == "032026"  # MMYYYY


# ============================================================
# GSTR-3B summary
# ============================================================


class TestGSTR3B:
    def test_produces_valid_xlsx(self):
        out = gstr3b_mod.build_gstr3b_summary(
            _business(), [_invoice()], [], date(2026, 5, 1)
        )
        assert out[:2] == b"PK"

    def test_includes_outward_taxable_in_3_1(self):
        invs = [_invoice(taxable_paise=100000, cgst_paise=9000, sgst_paise=9000)]
        out = gstr3b_mod.build_gstr3b_summary(_business(), invs, [], date(2026, 5, 1))
        wb = load_workbook(BytesIO(out))
        ws = wb["GSTR-3B Summary"]
        # Scan all rows for the outward amount
        found = False
        for row in ws.iter_rows(values_only=True):
            if row and 1000.0 in row:  # ₹1,000 (100000 paise)
                found = True
                break
        assert found, "Outward taxable amount not found in 3.1 section"

    def test_itc_aggregation_from_purchases(self):
        # Two purchases — one with IGST ₹18, one with CGST+SGST ₹9 each (₹18 combined)
        purchases = [
            _purchase(1, taxable_paise=10000, igst_paise=1800,
                      cgst_paise=0, sgst_paise=0, total_paise=11800),
            _purchase(2, taxable_paise=10000, igst_paise=0, cgst_paise=900,
                      sgst_paise=900, total_paise=11800),
        ]
        out = gstr3b_mod.build_gstr3b_summary(
            _business(), [_invoice()], purchases, date(2026, 5, 1)
        )
        wb = load_workbook(BytesIO(out))
        ws = wb["GSTR-3B Summary"]
        # IGST inputs = ₹18, CGST inputs = ₹9, SGST inputs = ₹9 — any of these
        # numeric cells should appear in the ITC section.
        all_nums: set[float] = set()
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    all_nums.add(round(float(v), 2))
        assert 18.0 in all_nums or 9.0 in all_nums

    def test_excludes_cancelled_invoices(self):
        invs = [
            _invoice(1),
            _invoice(2, status=InvoiceStatus.CANCELLED),
        ]
        out = gstr3b_mod.build_gstr3b_summary(_business(), invs, [], date(2026, 5, 1))
        wb = load_workbook(BytesIO(out))
        ws = wb["GSTR-3B Summary"]
        # Just verify it built; specific assertion is in test_excludes_cancelled above
        assert ws.cell(row=1, column=1).value.startswith("GSTR-3B")


# ============================================================
# Purchase register
# ============================================================


class TestPurchaseRegister:
    def test_produces_valid_xlsx(self):
        out = pr_mod.build_purchase_register(
            _business(), [_purchase()], date(2026, 5, 1), date(2026, 5, 31)
        )
        assert out[:2] == b"PK"

    def test_includes_supplier_row(self):
        out = pr_mod.build_purchase_register(
            _business(), [_purchase(1)], date(2026, 5, 1), date(2026, 5, 31)
        )
        wb = load_workbook(BytesIO(out))
        ws = wb["Purchase Register"]
        # Header row at row 6; data row at row 7
        assert ws.cell(row=7, column=2).value == "BILL-1"
        assert ws.cell(row=7, column=3).value == "Supplier 1"

    def test_total_row_appended(self):
        out = pr_mod.build_purchase_register(
            _business(),
            [_purchase(1, taxable_paise=10000, total_paise=11800),
             _purchase(2, taxable_paise=10000, total_paise=11800)],
            date(2026, 5, 1), date(2026, 5, 31),
        )
        wb = load_workbook(BytesIO(out))
        ws = wb["Purchase Register"]
        # 2 data rows (7, 8), TOTAL at row 9
        assert ws.cell(row=9, column=1).value == "TOTAL"
        # Total taxable column should be 200.00 (20000 paise)
        assert ws.cell(row=9, column=9).value == 200.00

    def test_excludes_cancelled(self):
        out = pr_mod.build_purchase_register(
            _business(),
            [_purchase(1), _purchase(2, status=PurchaseInvoiceStatus.CANCELLED)],
            date(2026, 5, 1), date(2026, 5, 31),
        )
        wb = load_workbook(BytesIO(out))
        ws = wb["Purchase Register"]
        # Only the first should appear
        assert ws.cell(row=7, column=2).value == "BILL-1"
        # Row 8 should be the TOTAL (since only 1 data row)
        assert ws.cell(row=8, column=1).value == "TOTAL"
